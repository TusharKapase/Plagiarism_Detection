import os
from datetime import datetime
import pandas as pd

# PDF Libraries
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


# ==========================================================
# FORMAT STUDENT NAME
# ==========================================================
def _format_name(filename):
    """
    Example:
    23510002_A1_-_23510002_MANE_BHUMIKA_SACHIN_1.py

    Output:
    MANE BHUMIKA SACHIN (23510002)
    """
    try:
        name = filename.split(".")[0]
        parts = name.split("_")

        prn = parts[0]
        student_name = " ".join(parts[4:-1])

        return f"{student_name} ({prn})"

    except:
        return filename


# ==========================================================
# EXCEL REPORT
# ==========================================================

def generate_excel_report(result_data, output_folder="reports"):
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    os.makedirs(output_folder, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(output_folder, f"Plagiarism_Report_{timestamp}.xlsx")

    rows = []

    for entry in result_data["details"]:
        rows.append({
            "Student 1": _format_name(entry["file1"]),
            "Student 2": _format_name(entry["file2"]),
            "Similarity %": round(entry["similarity"], 2),
            "Status": (
                "HIGH RISK" if entry["similarity"] >= 70
                else "MEDIUM" if entry["similarity"] >= 40
                else "LOW"
            )
        })

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Report", startrow=4, index=False)

        wb = writer.book
        ws = writer.sheets["Report"]

        # ==================================================
        # TITLE
        # ==================================================
        ws.merge_cells("A1:D1")
        ws["A1"] = "PLAGIARISM DETECTION REPORT"
        ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:D2")
        ws["A2"] = f"Generated on {datetime.now().strftime('%d-%m-%Y %I:%M %p')}"
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:D3")
        ws["A3"] = f"Average Similarity : {result_data['average_score']}%"
        ws["A3"].alignment = Alignment(horizontal="center")
        ws["A3"].font = Font(bold=True)

        # ==================================================
        # HEADER DESIGN
        # ==================================================
        header_fill = PatternFill("solid", fgColor="4472C4")

        for cell in ws[5]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # ==================================================
        # ROW COLORS
        # ==================================================
        thin = Side(style="thin", color="CCCCCC")

        for row in range(6, ws.max_row + 1):
            sim = ws[f"C{row}"].value

            for col in range(1, 5):
                c = ws.cell(row=row, column=col)
                c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                c.alignment = Alignment(vertical="center")

            if sim >= 70:
                fill = "FFC7CE"   # red
            elif sim >= 40:
                fill = "FFF2CC"   # yellow
            else:
                fill = "C6EFCE"   # green

            ws[f"C{row}"].fill = PatternFill("solid", fgColor=fill)
            ws[f"D{row}"].fill = PatternFill("solid", fgColor=fill)

        # ==================================================
        # AUTO WIDTH
        # ==================================================
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 4

    return report_path


# ==========================================================
# PDF REPORT
# ==========================================================


def generate_pdf_report(result_data, output_folder="reports"):

    os.makedirs(output_folder, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(output_folder, f"Plagiarism_Report_{timestamp}.pdf")

    doc = SimpleDocTemplate(
        report_path,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'title',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor("#1F4E78"),
        alignment=1,
        spaceAfter=20
    )

    subtitle_style = ParagraphStyle(
        'sub',
        parent=styles['Normal'],
        fontSize=12,
        alignment=1,
        textColor=colors.grey
    )

    story = []

    # ==================================================
    # COVER PAGE
    # ==================================================
    story.append(Spacer(1, 80))
    story.append(Paragraph("PLAGIARISM DETECTION REPORT", title_style))
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"Generated on {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
        subtitle_style
    ))
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"<b>Average Similarity:</b> {result_data['average_score']}%",
        subtitle_style
    ))

    story.append(PageBreak())

    # ==================================================
    # TABLE DATA
    # ==================================================
    data = [["Student 1", "Student 2", "Similarity %", "Risk"]]

    for entry in result_data["details"]:
        sim = round(entry["similarity"], 2)

        if sim >= 70:
            risk = "HIGH"
        elif sim >= 40:
            risk = "MEDIUM"
        else:
            risk = "LOW"

        data.append([
            _format_name(entry["file1"]),
            _format_name(entry["file2"]),
            f"{sim}%",
            risk
        ])

    table = Table(data, colWidths=[180, 180, 80, 70])

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
            colors.whitesmoke,
            colors.lightgrey
        ])
    ])

    # Risk color rows
    for i in range(1, len(data)):
        risk = data[i][3]

        if risk == "HIGH":
            style.add("TEXTCOLOR", (3, i), (3, i), colors.red)

        elif risk == "MEDIUM":
            style.add("TEXTCOLOR", (3, i), (3, i), colors.orange)

        else:
            style.add("TEXTCOLOR", (3, i), (3, i), colors.green)

    table.setStyle(style)

    story.append(Paragraph("Detailed Comparison Report", styles["Heading2"]))
    story.append(Spacer(1, 10))
    story.append(table)

    doc.build(story)

    return report_path










# import os
# from datetime import datetime
# import pandas as pd
# from reportlab.lib.pagesizes import letter
# from reportlab.pdfgen import canvas
# from reportlab.lib.units import inch
# from textwrap import wrap


# def _format_name(filename):
#     """
#     Converts:
#     23510002_A1_-_23510002_MANE_BHUMIKA_SACHIN_1.py
#     →
#     MANE_BHUMIKA_SACHIN_23510002
#     """
#     name = filename.replace(".py", "")
#     parts = name.split("_")

#     prn = parts[0]
#     person_name = "_".join(parts[4:-1])

#     return f"{person_name}_{prn}"







# def generate_excel_report(result_data, output_folder="reports"):
#     os.makedirs(output_folder, exist_ok=True)
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     report_path = os.path.join(output_folder, f"report_{timestamp}.xlsx")

#     df = pd.DataFrame(result_data["details"])
#     df.to_excel(report_path, index=False, engine="openpyxl")

#     return report_path


# def generate_pdf_report(result_data, output_folder="reports"):
#     os.makedirs(output_folder, exist_ok=True)
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     report_path = os.path.join(output_folder, f"report_{timestamp}.pdf")

#     c = canvas.Canvas(report_path, pagesize=letter)
#     width, height = letter

#     # Title
#     c.setFont("Helvetica-Bold", 16)
#     c.drawString(50, height - 50, "Plagiarism Detection Report")

#     # Average score
#     c.setFont("Helvetica", 12)
#     c.drawString(50, height - 80, f"Average Similarity: {result_data['average_score']}%")

#     y = height - 120
#     c.setFont("Helvetica", 10)

#     for entry in result_data["details"]:
#         name1 = _format_name(entry["file1"])
#         name2 = _format_name(entry["file2"])
#         similarity = f"{entry['similarity']}%"

#         line = f"{name1}  vs  {name2}  →  {similarity}"

#         wrapped_lines = wrap(line, 90)

#         for wline in wrapped_lines:
#             if y < 60:
#                 c.showPage()
#                 c.setFont("Helvetica", 10)
#                 y = height - 50

#             c.drawString(50, y, wline)
#             y -= 14

#         y -= 10  # extra space between entries

#     c.save()
#     return report_path
